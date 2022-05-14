import openpyxl as op
from random import randint
from time import sleep
from Functions import mostrar_pergunta_respostas, verificar_resposta, caucular_rank
from Functions import Menu_Quiz
def cadastro():
      #cadastro de usuarios 
    usuario = []
    #coleta de dados
    while True:
        nome_valido = 0
        #coleta nomes
        nome = input('\nCadastro de usuario: ').capitalize().strip()
        for n in sheet_cadastro['A']:
            if nome == n.value:
                print('usuario já registrado!')
                nome_valido += 1
                break
        if nome_valido > 0 or len(nome) < 3:
            continue
        else:
            usuario.append(nome)
            while True:
                #coleta de senhas
                senha = input('\nCadastro de senha: ').strip()
                if len(senha) < 6:
                    print('A senha deve conter no minimo 6 caracteres ')
                    continue
                else:
                    usuario.append(senha)
                    break
        break
    return usuario


def registro(sheet, usuario, book):
    sheet.append(usuario)
    book.save('C:\\Users\\asss1\\OneDrive\\Área de Trabalho\\Projeto do jogo\\Quiz_Game\\Pasta1.xlsx')

# Main Program
quiz_book = op.load_workbook('C:\\Users\\asss1\\OneDrive\\Área de Trabalho\\Projeto do jogo\\Quiz_Game\\Pasta1.xlsx')
sheet_quiz = quiz_book['Quiz']
sheet_cadastro = quiz_book['Cadastro']

# Debug linhas perguntas
# linhas_totais = sheet_quiz.max_row
# print(linhas_totais)

# Apresentando o quiz
Menu_Quiz.Menu("QUIZ DIVINO")

usuario = cadastro()
registro(sheet_cadastro, usuario, quiz_book)
    



""" # Determina o maximo de perguntas que existe na planilha
linhas = sheet_quiz.max_row - 1
Perguntas_Disponiveis = []
pontos = perguntas_respondidas = 0

# Loop onde o jogo so termina se o jogador quiser, ou não tiver mais nenhuma pergunta disponivel
while len(Perguntas_Disponiveis) != linhas:
    # Escolhe uma pergunta "aleatoriamente"
    num_perg = randint(1, linhas)
    # Se a pergunta não ja tiver sido escolhida, o programa procede com o jogo
    if num_perg not in Perguntas_Disponiveis:
        Perguntas_Disponiveis.append(num_perg)
        mostrar_pergunta_respostas(sheet_quiz, num_perg)
        while True:
            resposta = input('Qual a alternativa correta? ').upper().strip()
            # Verificando se a resposta do usuario é válida
            if resposta not in 'ABC' or resposta == '':
                print('\033[31m!! Digite uma alternativa válida !!\033[m')
                continue
            else:
                # Verifica a resposta do usuario é correda ou não
                verificador = verificar_resposta(sheet_quiz, num_perg, resposta)
                if verificador > 0:
                    if verificador == 2:
                        pontos += 2
                        perguntas_respondidas += 1
                    else:
                        pontos += 1
                        perguntas_respondidas += 1
                else:
                    perguntas_respondidas += 1
                break

    # Após a primeira onda de perguntas, o programa irá perguntar se o jogador irá querer continuar.
    loop_continue = input('Deseja tentar mais uma pergunta?[S/N] ').strip().upper()
    while loop_continue not in 'SN' or loop_continue == '':
        print('\033[31m!! Digite uma alternativa válida !!\033[m')
        loop_continue = input('Deseja tentar mais uma pergunta?[S/N] ').strip().upper()

    # Caso o usuario não queira, o programa irá salvar os resultados da partida e então irá parar
    if loop_continue == 'N':
        print('Finalizando o QUIZ...')
        sleep(1)
        print('Registrando jogador...')
        rank = caucular_rank(pontos, perguntas_respondidas)
        sleep(1)
        print('QUIZ FINALIZADO.')
        break

    else:
        print('Criando uma nova pergunta...')
        sleep(1)
 """
