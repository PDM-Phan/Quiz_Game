import openpyxl as op
from random import randint
from Funções_para_Planilha import mostrar_planilha, mostrar_pergunta_respostas

quiz_planilha = op.load_workbook('Pasta1.xlsx')

# Visualizando os sheets de uma planilha
#   print(quiz_planilha.sheetnames)
# Printando conteudo desse sheet
#   sheet_quiz = quiz_planilha['Quiz']
#   quiz_max_row = sheet_quiz.max_row
#   quiz_max_col = sheet_quiz.max_column
#
#   mostrar_planilha(sheet_quiz, quiz_max_col, quiz_max_row)

sheet_quiz = quiz_planilha['Quiz']
limite_perguntas = []
while len(limite_perguntas) != 5:
    num_perg = randint(1, 5)
    if num_perg not in limite_perguntas:
        limite_perguntas.append(num_perg)
        mostrar_pergunta_respostas(sheet_quiz, num_perg)
